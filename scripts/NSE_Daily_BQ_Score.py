import yfinance as yf
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from google.cloud import bigquery
from google.api_core.exceptions import NotFound
from openpyxl import Workbook, load_workbook
import csv
import os
import pytz
import time
import logging
import pandas as pd

# Define IST timezone
IST = pytz.timezone('Asia/Kolkata')

# Current IST datetime
ist_now = datetime.now(IST)

# Extract the date part from IST datetime
ist_date = ist_now.date()

# Generate log and CSV file names 
log_filename = f"log_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}_score1.txt"
master_log_filename = f"Log_Master_NSE_BigQuery_test_dump_score1.txt"
csv_filename = f"NSE_Stock_Master_BQ_test_dump_score1.csv"  # Append data for the same day
csv_filename_daily = f"NSE_Stock_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}_test_dump_score1.csv"  # Append data for the same day
excel_filename = f"NSE_Stock_Master_DataLake_test_dump_score1.xlsx"  # Excel file for today

# Paths for logs, CSV, and Excel
MASTER_LOG_FILE_PATH = os.path.join("logs", master_log_filename)
LOG_FILE_PATH = os.path.join("logs", log_filename)
MASTER_CSV_FILE_PATH = os.path.join("csv", csv_filename)
Daily_CSV_FILE_PATH = os.path.join("csv", csv_filename_daily)
EXCEL_FILE_PATH = os.path.join("excel", excel_filename)

# Ensure directories exist
os.makedirs("master_log", exist_ok=True)
os.makedirs("logs", exist_ok=True)
os.makedirs("csv", exist_ok=True)
os.makedirs("excel", exist_ok=True)

# Log function
def log_message(message):
    """Log messages to a file and print to console."""
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE_PATH, "a") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")
    with open(MASTER_LOG_FILE_PATH, "a") as master_log_file:
        master_log_file.write(f"[{timestamp}] {message}\n")
    print(f"[{timestamp}] {message}")

# Authenticate using the same service_account.json for both BigQuery and Google Sheets
SERVICE_ACCOUNT_FILE = "service_account.json"

# Google Sheets authentication
gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)


# Open Google Spreadsheet
spreadsheet = gc.open('NSE_symbol')  # Replace with your Google Sheet name
#source_worksheet = spreadsheet.worksheet('symbol')  # Replace with your sheet name
source_worksheet = spreadsheet.worksheet('symbol_test')  # Test sheet name

# Fetch all stock symbols from the first column
symbols = source_worksheet.col_values(1)[1:]  # Skip header row
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Define BigQuery dataset and table with the project ID
PROJECT_ID = "stockautomation-442015"  # Replace with your project ID
BQ_DATASET = "nse_stock_test_score3"  # Replace with your dataset name
BQ_TABLE = f"{PROJECT_ID}.{BQ_DATASET}.daily_nse_stock_data_test1"  # Fully-qualified table name

# Define schema for BigQuery table
headers = [
    "industry", "sector", "fullTimeEmployees", "auditRisk", "boardRisk",
    "compensationRisk", "shareHolderRightsRisk", "overallRisk", "maxAge", "priceHint",
    "previousClose", "open", "dayLow", "dayHigh", "regularMarketPreviousClose",
    "regularMarketOpen", "regularMarketDayLow", "regularMarketDayHigh", "dividendRate",
    "dividendYield", "exDividendDate", "payoutRatio", "fiveYearAvgDividendYield", "beta",
    "trailingPE", "forwardPE", "volume", "regularMarketVolume", "averageVolume",
    "averageVolume10days", "averageDailyVolume10Day", "marketCap", "fiftyTwoWeekLow",
    "fiftyTwoWeekHigh", "priceToSalesTrailing12Months", "fiftyDayAverage",
    "twoHundredDayAverage", "trailingAnnualDividendRate", "trailingAnnualDividendYield",
    "enterpriseValue", "profitMargins", "floatShares", "sharesOutstanding",
    "heldPercentInsiders", "heldPercentInstitutions", "impliedSharesOutstanding",
    "bookValue", "priceToBook", "earningsQuarterlyGrowth", "trailingEps", "forwardEps",
    "52WeekChange", "lastDividendValue", "lastDividendDate", "exchange", "quoteType",
    "symbol", "shortName", "longName", "currentPrice", "targetHighPrice", "targetLowPrice",
    "targetMeanPrice", "targetMedianPrice", "recommendationMean", "recommendationKey",
    "numberOfAnalystOpinions", "totalCash", "totalCashPerShare", "ebitda", "totalDebt",
    "quickRatio", "currentRatio", "totalRevenue", "debtToEquity", "revenuePerShare",
    "returnOnAssets", "returnOnEquity", "freeCashflow", "operatingCashflow",
    "earningsGrowth", "revenueGrowth", "grossMargins", "ebitdaMargins", "operatingMargins"
]

# Define a data type mapping for headers
data_type_map = {
    "industry": "STRING",
    "sector": "STRING",
    "fullTimeEmployees": "FLOAT",  # Integer field
    "auditRisk": "FLOAT",
    "boardRisk": "FLOAT",
    "compensationRisk": "FLOAT",
    "shareHolderRightsRisk": "FLOAT",
    "overallRisk": "FLOAT",
    "maxAge": "INTEGER",
    "priceHint": "INTEGER",
    "previousClose": "FLOAT",
    "open": "FLOAT",
    "dayLow": "FLOAT",
    "dayHigh": "FLOAT",
    "regularMarketPreviousClose": "FLOAT",
    "regularMarketOpen": "FLOAT",
    "regularMarketDayLow": "FLOAT",
    "regularMarketDayHigh": "FLOAT",
    "dividendRate": "FLOAT",
    "dividendYield": "FLOAT",
    "exDividendDate": "DATE",
    "payoutRatio": "FLOAT",
    "fiveYearAvgDividendYield": "FLOAT",
    "beta": "FLOAT",
    "trailingPE": "FLOAT",
    "forwardPE": "FLOAT",
    "volume": "INTEGER",
    "regularMarketVolume": "INTEGER",
    "averageVolume": "INTEGER",
    "averageVolume10days": "INTEGER",
    "averageDailyVolume10Day": "INTEGER",
    "marketCap": "INTEGER",
    "fiftyTwoWeekLow": "FLOAT",
    "fiftyTwoWeekHigh": "FLOAT",
    "priceToSalesTrailing12Months": "FLOAT",
    "fiftyDayAverage": "FLOAT",
    "twoHundredDayAverage": "FLOAT",
    "trailingAnnualDividendRate": "FLOAT",
    "trailingAnnualDividendYield": "FLOAT",
    "enterpriseValue": "INTEGER",
    "profitMargins": "FLOAT",
    "floatShares": "INTEGER",
    "sharesOutstanding": "INTEGER",
    "heldPercentInsiders": "FLOAT",
    "heldPercentInstitutions": "FLOAT",
    "impliedSharesOutstanding": "INTEGER",
    "bookValue": "FLOAT",
    "priceToBook": "FLOAT",
    "earningsQuarterlyGrowth": "FLOAT",
    "trailingEps": "FLOAT",
    "forwardEps": "FLOAT",
    "52WeekChange": "FLOAT",
    "lastDividendValue": "FLOAT",
    "lastDividendDate": "DATE",
    "exchange": "STRING",
    "quoteType": "STRING",
    "symbol": "STRING",
    "shortName": "STRING",
    "longName": "STRING",
    "currentPrice": "FLOAT",
    "targetHighPrice": "FLOAT",
    "targetLowPrice": "FLOAT",
    "targetMeanPrice": "FLOAT",
    "targetMedianPrice": "FLOAT",
    "recommendationMean": "FLOAT",
    "recommendationKey": "STRING",
    "numberOfAnalystOpinions": "INTEGER",
    "totalCash": "INTEGER",
    "totalCashPerShare": "FLOAT",
    "ebitda": "INTEGER",
    "totalDebt": "INTEGER",
    "quickRatio": "FLOAT",
    "currentRatio": "FLOAT",
    "totalRevenue": "INTEGER",
    "debtToEquity": "FLOAT",
    "revenuePerShare": "FLOAT",
    "returnOnAssets": "FLOAT",
    "returnOnEquity": "FLOAT",
    "freeCashflow": "INTEGER",
    "operatingCashflow": "INTEGER",
    "earningsGrowth": "FLOAT",
    "revenueGrowth": "FLOAT",
    "grossMargins": "FLOAT",
    "ebitdaMargins": "FLOAT",
    "operatingMargins": "FLOAT",
    "Today_Growth": "FLOAT",
    "Calculated_Score": "FLOAT",
    "Score_Recommendation": "STRING",
    "Conservative_Invs_Recom": "STRING",
    "Conservative_Invs_Reson": "STRING",
    "Growth_Invs_Recom": "STRING",
    "Growth_Invs_Reson": "STRING",
    "Momentum_Invs_Recom": "STRING",
    "Momentum_Invs_Reson": "STRING",
}

ROW_COUNTER_FILE = "master/nse_row_counter.txt"

# Initialize row_insert_order
def initialize_row_counter():
    if not os.path.exists(ROW_COUNTER_FILE):
        with open(ROW_COUNTER_FILE, "w") as f:
            f.write("1")  # Start counter at 0

def get_current_row_counter():
    with open(ROW_COUNTER_FILE, "r") as f:
        return int(f.read().strip())

def update_row_counter(new_value):
    with open(ROW_COUNTER_FILE, "w") as f:
        f.write(str(new_value))

# Initialize the row counter if not already done
initialize_row_counter()


# Add "Previous Day Date" to headers
# PREVIOUS_DAY_DATE = (ist_date - timedelta(days=1)).strftime('%Y-%m-%d')
PREVIOUS_DAY_DATETIME = (ist_now - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
headers_with_date = ["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers

score_headers = ["Today_Growth", "Calculated_Score", "Score_Recommendation", "Conservative_Invs_Recom", "Conservative_Invs_Reson","Growth_Invs_Recom", "Growth_Invs_Reson","Momentum_Invs_Recom", "Momentum_Invs_Reson"]

def ensure_dataset_exists():
    try:
        bq_client.get_dataset(BQ_DATASET)
        log_message(f"Dataset '{BQ_DATASET}' exists.")
    except NotFound:
        dataset = bigquery.Dataset(f"{PROJECT_ID}.{BQ_DATASET}")
        bq_client.create_dataset(dataset)
        log_message(f"Created dataset '{BQ_DATASET}'.")

def ensure_table_exists():
    try:
        # Check if the table already exists
        table = bq_client.get_table(BQ_TABLE)
        log_message(f"Table '{BQ_TABLE}' already exists.")
    except NotFound:
        # Table does not exist, create it
        # Build the schema dynamically
        schema = [bigquery.SchemaField("row_insert_order", "INTEGER"), bigquery.SchemaField("PreviousDayDate", "DATETIME"), bigquery.SchemaField("Symbol_Input", "STRING"),] + [
                bigquery.SchemaField(header, data_type_map.get(header, "STRING"))
                for header in headers
                ] + [ bigquery.SchemaField(header, data_type_map.get(header, "STRING")) for header in score_headers ]
        
        table = bigquery.Table(BQ_TABLE, schema=schema)
        bq_client.create_table(table)
        log_message(f"Created table '{BQ_TABLE}'.")
    except Exception as e:
        log_message(f"Error ensuring table exists: {e}")

def append_to_csv(data_row):
    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(MASTER_CSV_FILE_PATH)  # Check if file exists

    with open(MASTER_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header row
            log_message(f"Header added to CSV file: {MASTER_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to CSV file: {MASTER_CSV_FILE_PATH}")

    """Append a row of data to the CSV file, adding the header only if it's a new file."""
    write_header = not os.path.exists(Daily_CSV_FILE_PATH)  # Check if file exists

    with open(Daily_CSV_FILE_PATH, mode="a", newline="") as csv_file:
        writer = csv.writer(csv_file)
        if write_header:
            writer.writerow(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header row
            log_message(f"Header added to CSV file: {Daily_CSV_FILE_PATH}")
        writer.writerow(data_row)
        log_message(f"Appended data to CSV file: {Daily_CSV_FILE_PATH}")

def append_to_excel(data_row):
    """Append data to an Excel sheet, creating a new sheet for the day."""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
        else:
            workbook = Workbook()

        sheet_name = f"NSE_{ist_date}"
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.append(["row_insert_order", "PreviousDayDate", "Symbol_Input"] + headers + score_headers)  # Add header
        else:
            sheet = workbook[sheet_name]

        sheet.append(data_row)
        
        # Freeze the first row and third column
        sheet.freeze_panes = 'D2'  # Freeze everything above row 2 and to the left of column C

        workbook.save(EXCEL_FILE_PATH)
        log_message(f"Data appended to Excel file: {EXCEL_FILE_PATH}")
    except Exception as e:
        log_message(f"Error saving to Excel: {e}")

def validate_input(value, min_val=None):
    if value is None or pd.isna(value) or not isinstance(value, (int, float)):
        return None
    if min_val is not None and value < min_val:
        return None
    return value
        
def calculate_individual_scores(pe, dividend_yield, earnings_growth):

    dividend_yield = dividend_yield * 100

    # P/E Ratio Scoring Logic (Lower is better)
    if pe is not None:
        if pe <= 10:
            pe_score = 5
        elif pe <= 20:
            pe_score = 4
        elif pe <= 30:
            pe_score = 3
        elif pe <= 50:
            pe_score = 2
        else:
            pe_score = 1
    else:
        pe_score = 0

    # Dividend Yield Scoring Logic (Higher is better)
    if dividend_yield is not None:
        if dividend_yield > 4:
            dividend_score = 5
        elif dividend_yield > 3:
            dividend_score = 4
        elif dividend_yield > 2:
            dividend_score = 3
        elif dividend_yield > 1:
            dividend_score = 2
        else:
            dividend_score = 1
    else:
        dividend_score = 0

    # Earnings Growth Scoring Logic (Higher is better)
    if earnings_growth is not None:
        if earnings_growth > 20:
            earnings_growth_score = 5
        elif earnings_growth > 10:
            earnings_growth_score = 4
        elif earnings_growth > 5:
            earnings_growth_score = 3
        elif earnings_growth > 0:
            earnings_growth_score = 2
        else:
            earnings_growth_score = 1
    else:
        earnings_growth_score = 0

    # If any of the scores are invalid, return "None"
    if pe_score == 0 or dividend_score == 0 or earnings_growth_score == 0:
        return None

    # Weighted total score (scaled to 1-5)
    total_score = (pe_score * 0.4) + (dividend_score * 0.3) + (earnings_growth_score * 0.3)
    total_score = round(total_score, 1)
    
    # Assign Calculated Recommendation
    if total_score <= 1.5:
        recommendation = "Strong Buy"
    elif total_score <= 2.5:
        recommendation = "Buy"
    elif total_score <= 3.5:
        recommendation = "Hold"
    elif total_score <= 4.5:
        recommendation = "Underperform"
    else:
        recommendation = "Sell"

    return total_score, recommendation
    #return round(total_score, 1)

def analyze_stock_with_profiles(info):
    recommendations = []
    
    try:
        # Extract relevant fields
        beta = info.get('beta', 'N/A')
        pe_ratio = info.get('trailingPE', 'N/A')
        forward_pe = info.get('forwardPE', 'N/A')
        dividend_yield = info.get('dividendYield', 'N/A')
        price_to_book = info.get('priceToBook', 'N/A')
        profit_margins = info.get('profitMargins', 'N/A')
        revenue_growth = info.get('revenueGrowth', 'N/A')
        high_52w = info.get('fiftyTwoWeekHigh', 'N/A')
        low_52w = info.get('fiftyTwoWeekLow', 'N/A')
        recommendation_mean = info.get('recommendationMean', 'N/A')
        current_price = info.get('currentPrice', 'N/A')

        # 1. Conservative Investor (Low Risk, Income-Focused)
        if beta != 'N/A' and beta < 1:
            conservative_reason = "Low Beta (less volatile than the market)"
        else:
            conservative_reason = "High Beta (more volatile)"
        
        if dividend_yield != 'N/A' and dividend_yield > 0.03:
            conservative_reason += ",\n Pays a good dividend (>3%)"
        
        if price_to_book != 'N/A' and price_to_book < 1:
            conservative_reason += ",\n Price-to-Book ratio (<1) indicates undervalued assets"
        elif price_to_book != 'N/A' and price_to_book < 2:
            conservative_reason += ",\n Price-to-Book ratio (<2) indicates potential for growth"
        
        recommendations.append({
            "Cal_Investment_Profile": "Conservative Investor",
            "Cal_Recommendation": "Buy" if dividend_yield != 'N/A' and dividend_yield > 0.03 else "Hold",
            "Cal_Reason": conservative_reason
        })

        # 2. Growth Investor (Focus on High Growth)
        growth_reason = []
        if forward_pe != 'N/A' and forward_pe < 20:
            growth_reason.append("Low Forward P/E (<20) indicates growth potential")
        if revenue_growth != 'N/A' and revenue_growth > 0.1:
            growth_reason.append("Strong Revenue Growth (>10%)")
        if profit_margins != 'N/A' and profit_margins > 0.2:
            growth_reason.append("Highly Profitable with margins > 20%")
        
        if growth_reason:
            recommendations.append({
                "Cal_Investment_Profile": "Growth Investor",
                "Cal_Recommendation": "Buy" if forward_pe != 'N/A' and forward_pe < 20 else "Hold",
                "Cal_Reason": ",\n ".join(growth_reason)
            })

        # 3. Momentum Investor (Focus on Recent Trends)
        momentum_reason = []
        if high_52w != 'N/A' and low_52w != 'N/A':
            price_position = (current_price - low_52w) / (high_52w - low_52w)
            if price_position > 0.75:
                momentum_reason.append("Trading near its 52-week high (bullish momentum)")
            elif price_position < 0.25:
                momentum_reason.append("Trading near its 52-week low (bearish momentum)")
       
        recommendations.append({
            "Cal_Investment_Profile": "Momentum Investor",
            "Cal_Recommendation": "Buy" if price_position > 0.75 else "Hold",
            "Cal_Reason": ",\n ".join(momentum_reason) if momentum_reason else "No strong momentum"
        })

    except Exception as e:
        recommendations.append({
            "Cal_Investment_Profile": "Error",
            "Cal_Recommendation": "",
            "Cal_Reason": ""
        })
    
    return recommendations
    
    
def fetch_and_update_stock_data(symbol):
    try:
        # Read the current row counter
        current_counter = get_current_row_counter()

        log_message(f"Life count: {current_counter} Fetching data for {symbol} ...")
        stock = yf.Ticker(symbol)
        info = stock.info
        
        # Safely access data with default values
        pe_ratio = info.get('trailingPE', 0)
        dividend_yield = info.get('dividendYield', 0)
        earnings_growth = info.get('earningsQuarterlyGrowth', 0)        

        # Calculate score (assuming the `calculate_individual_scores` function is defined)
        score, score_recommendation = calculate_individual_scores(pe_ratio, dividend_yield, earnings_growth)

        cal_recom = analyze_stock_with_profiles(info)

        current_price = info.get('currentPrice', 'N/A')
        previous_close = info.get('previousClose', 'N/A')

        # Calculate today's growth percentage
        if current_price != 'N/A' and previous_close != 'N/A' and previous_close != 0:
            today_growth_percentage = ((current_price - previous_close) / previous_close) * 100
            today_growth_percentage = round(today_growth_percentage, 2)
        else:
            today_growth_percentage = 'N/A'

        # PREVIOUS_DAY_DATE = (ist_date - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
        PREVIOUS_DAY_DATETIME = (ist_now - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
        # Extract data and include the Previous Day Date
        info_row = [current_counter, PREVIOUS_DAY_DATETIME, symbol] + [info.get(key, '') for key in headers]  + [today_growth_percentage, score, score_recommendation]

        for recom in cal_recom:
            #info_row.append(recom.get("Cal_Investment_Profile", ""))
            info_row.append(recom.get("Cal_Recommendation", ""))
            info_row.append(recom.get("Cal_Reason", ""))
    
        # Increment row_insert_order for the next row
        current_counter += 1
        update_row_counter(current_counter)
        
        # Append data to CSV and Excel
        append_to_csv(info_row)
        append_to_excel(info_row)
       
        return info_row
    except Exception as e:
        log_message(f"Error fetching data for {symbol}: {e}")
        return None

# Add process data
def preprocess_data(csv_file_path):
    """
    Preprocess the CSV file to ensure data types are correct based on the BigQuery schema.
    If incorrect types are detected, log the error and attempt to fix them.
    """

    processed_rows = []
    errors = []

    try:
        with open(csv_file_path, "r") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                processed_row = {}
                for key, value in row.items():
                    expected_type = data_type_map.get(key, "STRING")
                    try:
                        if expected_type == "STRING":
                            processed_row[key] = value.strip() if value else None  # Handle empty strings as None
                        elif expected_type == "INTEGER":
                            processed_row[key] = int(value) if value else None
                        elif expected_type == "FLOAT":
                            try:
                                # Attempt to convert the value to a float
                                processed_row[key] = float(value)
                            except (ValueError, TypeError):
                                # If conversion fails, set it to None
                                processed_row[key] = None
                        elif expected_type == "DATETIME":
                            processed_row[key] = (
                                datetime.strptime(value, "%Y-%m-%d %H:%M:%S") 
                                if value 
                                else datetime(1990, 1, 1, 0, 0, 0)  # Default to '1990-01-01 00:00:00' if no value is provided
                            )
                        elif expected_type == "DATE":
                            try:
                                # Check if value is a Unix timestamp and convert it
                                if value.isdigit():
                                    processed_row[key] = datetime.fromtimestamp(int(value)).date()
                                else:
                                    # Parse date string in the format "YYYY-MM-DD"
                                    processed_row[key] = datetime.strptime(value, "%Y-%m-%d").date()
                            except Exception:
                                # Handle invalid or missing date values with a default date
                                processed_row[key] = datetime(1990, 1, 1).date()
                        else:  # STRING
                            processed_row[key] = value
                    except (ValueError, TypeError, KeyError) as ve:
                        errors.append(
                            f"Row {row_num}, Field '{key}' with value '{value}' failed conversion to {expected_type}: {ve}"
                        )
                        processed_row[key] = None  # Default to None on error
                        
                # Validate that the processed row has consistent column counts
                if len(processed_row) == len(row):
                    processed_rows.append(processed_row)
                else:
                    errors.append(f"Row {row_num} has inconsistent column counts.")
                    
                #processed_rows.append(processed_row)
    except Exception as e:
        log_message(f"Error reading or processing CSV file: {e}")
    
    # Log errors, if any
    if errors:
        log_message(f"Data type errors detected during preprocessing:\n" + "\n".join(errors))
    
    return processed_rows

def load_data_to_bigquery():
    """Load data from the preprocessed CSV file into BigQuery."""
    try:
        processed_data = preprocess_data(Daily_CSV_FILE_PATH)
        
        # Write processed data back to a temporary CSV for BigQuery loading
        temp_csv_path = "temp_processed.csv"
        with open(temp_csv_path, "w", newline="") as temp_csv:
            writer = csv.DictWriter(temp_csv, fieldnames=processed_data[0].keys())
            writer.writeheader()  # Write headers
            writer.writerows(processed_data)  # Write processed rows
        
        df = pd.read_csv(temp_csv_path)

        # Check if any rows have a different number of columns
        column_count = len(df.columns)
        print("column_count:", column_count)

        # Ensure all numeric columns are properly filled with numeric values
        for col in df.select_dtypes(include=["float64", "int64"]).columns:
            df[col] = df[col].fillna(0)  # Use a numeric placeholder like -1 for missing values in numeric columns

        # Ensure all string/object columns are properly filled with string values
        for col in df.select_dtypes(include=["object"]).columns:
            df[col] = df[col].fillna("")  # Use 'NULL' or any string value for missing values in string columns

        df.to_csv('temp_processed.csv', index=False)

        log_message(f"Start to load data to BigQuery from {temp_csv_path}.")

        # Load the processed data into BigQuery
        with open(temp_csv_path, "rb") as csv_file:
            job_config = bigquery.LoadJobConfig(
                source_format=bigquery.SourceFormat.CSV,
                skip_leading_rows=1,  # Skip header row
                write_disposition="WRITE_APPEND",  # Append data schema=schema
                autodetect=False,
                max_bad_records=500,  # Tolerate up to 50 bad rows
                ignore_unknown_values=True,  # Ignore unexpected columns
            )
            load_job = bq_client.load_table_from_file(
                csv_file, BQ_TABLE, job_config=job_config
            )
            load_job.result()  # Wait for the job to complete
            # Check for errors
            if load_job.errors:
                log_message(f"Errors encountered during loading: {job.errors}")
            else:
                log_message("Data loaded successfully, no errors.")
            log_message(f"Data successfully loaded to BigQuery from {temp_csv_path}.")
    except Exception as e:
        log_message(f"Error loading data to BigQuery: {e}")

# Process each symbol
processed_count = 0

# Process each symbol
for symbol in symbols:
    fetch_and_update_stock_data(symbol)
    processed_count += 1
    # Add a delay to avoid rate-limiting
    time.sleep(0.7)
    log_message(f"Processed {processed_count}/{len(symbols)} symbols.")

# BigQuery authentication
bq_client = bigquery.Client.from_service_account_json(SERVICE_ACCOUNT_FILE)

# Ensure dataset and table exist in BigQuery
ensure_dataset_exists()
ensure_table_exists()

# Load the data into BigQuery from the CSV file
load_data_to_bigquery()

log_message("Script execution completed.")
