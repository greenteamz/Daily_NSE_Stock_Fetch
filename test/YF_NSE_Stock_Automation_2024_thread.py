import yfinance as yf
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook

# List all the Excel files in the current directory
excel_files = [f for f in os.listdir() if f.endswith('.xlsx')]

# If there are no Excel files, print a message and exit
if not excel_files:
    print("No Excel files found in the directory.")
    exit()

# Print all Excel files found in the directory
print("Available Excel files:")
for i, file in enumerate(excel_files, start=1):
    print(f"{i}. {file}")

# Ask the user to choose an Excel file from the list
try:
    choice = int(input(f"Enter the number corresponding to the Excel file you want to use (1-{len(excel_files)}): "))
    selected_file = excel_files[choice - 1]
except (ValueError, IndexError):
    print("Invalid choice. Exiting the program.")
    exit()

# Load the selected Excel file using openpyxl
try:
    wb = load_workbook(selected_file)
except Exception as e:
    print(f"Error: Unable to load the Excel file. {e}")
    exit()

# Access the worksheet (tab) named 'symbol' in the workbook
if 'symbol' in wb.sheetnames:
    # Read 'symbol' sheet into a DataFrame
    source_worksheet = pd.read_excel(selected_file, sheet_name='symbol')
else:
    print("Error: 'symbol' sheet not found in the selected file.")
    exit()

# Fetch all stock symbols from the first column
symbols = source_worksheet['Symbol'].dropna().tolist()  # Drop NaN values and get the list of symbols
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Get today's date for the new sheet name
today = datetime.now().strftime('%d_%m_%Y')
sheet_name = f"Data_2024_to_{today}"

# Add a new sheet for the 2024 data
if sheet_name not in wb.sheetnames:
    sheet = wb.create_sheet(sheet_name)
    # Write the headers to the new sheet
    headers = ['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume', 'Ticker']
    sheet.append(headers)
else:
    sheet = wb[sheet_name]

# Prepare the text file for logging
output_text = "Fetched_Stock_Data.txt"
if not os.path.exists(output_text):
    with open(output_text, "w") as txt_file:
        txt_file.write(",".join(headers) + "\n")

# Define a function to fetch and save data for a single symbol
def fetch_and_save(symbol):
    try:
        print(f"Fetching data for {symbol}...")
        start_date = "2024-01-01"  # Start date for 2024 data
        end_date = datetime.now().strftime('%Y-%m-%d')  # End date is today
        data = yf.download(tickers=symbol, start=start_date, end=end_date, interval="1d", auto_adjust=True)
        
        if data.empty:
            print(f"No data found for {symbol}.")
            return

        # Add a column for the ticker symbol
        data['Ticker'] = symbol
        data.reset_index(inplace=True)

        # Append data to the Excel sheet row by row
        for _, row in data.iterrows():
            sheet.append(row.tolist())

        # Append data to the text file
        with open(output_text, "a") as txt_file:
            data.to_csv(txt_file, index=False, header=False, mode="a")

        print(f"Data for {symbol} written successfully.")

    except Exception as e:
        print(f"Error fetching data for {symbol}: {e}")

# Fetch and save data sequentially
for symbol in symbols:
    fetch_and_save(symbol)

# Save the workbook
try:
    wb.save(selected_file)
    print(f"Data successfully added to the worksheet titled '{sheet_name}'.")
except Exception as e:
    print(f"Error saving the workbook: {e}")

print(f"All data saved to {output_text} as CSV.")
