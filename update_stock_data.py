import yfinance as yf
from datetime import datetime, date
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from openpyxl.utils import get_column_letter
import os
import subprocess
import pytz
import time

# Define IST timezone
IST = pytz.timezone('Asia/Kolkata')

# Current IST datetime
ist_now = datetime.now(IST)

# Extract the date part from IST datetime
ist_date = ist_now.date()

# Generate log file name with date and time
log_filename = f"log_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
csv_filename = f"symbol_data_{ist_now.strftime('%Y-%m-%d_%H-%M-%S')}.csv"

# Set log and CSV file paths in the Git project
LOG_FILE_PATH = os.path.join("logs", log_filename)
CSV_FILE_PATH = os.path.join("csv", csv_filename)

# Ensure the log and CSV directories exist, create them if they don't
os.makedirs("logs", exist_ok=True)
os.makedirs("csv", exist_ok=True)

# Log message function
def log_message(message):
    """Write messages to log file and print them to console."""
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE_PATH, "a") as log_file:
        log_file.write(f"[{timestamp}] {message}\n")
    print(f"[{timestamp}] {message}")

# Authenticate Google Sheets API
gc = gspread.service_account(filename="service_account.json")

# Open the Google Spreadsheet (workbook) named 'NSE_symbol'
spreadsheet = gc.open('NSE_symbol')  # Replace 'NSE_symbol' with the actual workbook name

# Access the worksheet (tab) named 'symbol'
source_worksheet = spreadsheet.worksheet('symbol')  # Access the 'symbol' sheet

# Fetch all stock symbols from the first column
symbols = source_worksheet.col_values(1)[1:]  # Skip the header row

# Add '.NS' suffix to all symbols if not present
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Define Google Sheet name with date
base_sheet_name = f"NSE_{ist_date.strftime('%d_%m_%Y')}"
existing_sheet_names = [ws.title for ws in spreadsheet.worksheets()]

# Check and define the Google Sheet name
if base_sheet_name in existing_sheet_names:
    sheet_name = f"{base_sheet_name}_{ist_now.strftime('%H-%M-%S')}"
else:
    sheet_name = base_sheet_name

# Create a new Google Sheet
new_worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="1000", cols="100")

# Define the headers for the Google Sheet
headers = [
    "Market Cap", "industry", "sector", "fullTimeEmployees", "auditRisk", "boardRisk",
    "compensationRisk", "shareHolderRightsRisk", "overallRisk", "maxAge", "priceHint",
    "previousClose", "open", "dayLow", "dayHigh", "regularMarketPreviousClose",
    "regularMarketOpen", "regularMarketDayLow", "regularMarketDayHigh", "dividendRate",
    "dividendYield", "exDividendDate", "payoutRatio", "fiveYearAvgDividendYield", "beta",
    "trailingPE", "forwardPE", "volume", "regularMarketVolume", "averageVolume",
    "averageVolume10days", "averageDailyVolume10Day", "marketCap", "fiftyTwoWeekLow",
    "fiftyTwoWeekHigh", "priceToSalesTrailing12Months", "fiftyDayAverage",
    "twoHundredDayAverage", "trailingAnnualDividendRate", "trailingAnnualDividendYield",
    "enterpriseValue", "profitMargins", "floatShares", "sharesOutstanding",
    "heldPercentInsiders", "heldPercentInstitutions", "impliedSharesOutstanding",
    "bookValue", "priceToBook", "earningsQuarterlyGrowth", "trailingEps", "forwardEps",
    "52WeekChange", "lastDividendValue", "lastDividendDate", "exchange", "quoteType",
    "symbol", "shortName", "longName", "currentPrice", "targetHighPrice", "targetLowPrice",
    "targetMeanPrice", "targetMedianPrice", "recommendationMean", "recommendationKey",
    "numberOfAnalystOpinions", "totalCash", "totalCashPerShare", "ebitda", "totalDebt",
    "quickRatio", "currentRatio", "totalRevenue", "debtToEquity", "revenuePerShare",
    "returnOnAssets", "returnOnEquity", "freeCashflow", "operatingCashflow",
    "earningsGrowth", "revenueGrowth", "grossMargins", "ebitdaMargins", "operatingMargins"
]

# Set the headers in the Google Sheet
new_worksheet.append_row(['Symbol'] + headers)

# Write headers to CSV file
with open(CSV_FILE_PATH, "w") as csv_file:
    csv_file.write(",".join(['Symbol'] + headers) + "\n")

# Function to fetch and update data for a single symbol
def fetch_and_update_stock_data(symbol, worksheet, excel_sheet=None):
    try:
        log_message(f"Fetching data for {symbol}...")
        stock = yf.Ticker(symbol)
        info = stock.info

        # Extract selected fields
        info_row = [symbol] + [str(info.get(key, '')) for key in headers]
        worksheet.append_row(info_row)

        # Append row to CSV file
        with open(CSV_FILE_PATH, "a") as csv_file:
            csv_file.write(",".join(info_row) + "\n")

        # Update Excel sheet if provided
        if excel_sheet:
            excel_sheet.append(info_row)

        log_message(f"Successfully updated data for {symbol}.")
    except Exception as e:
        log_message(f"Error fetching data for {symbol}: {e}")

# Path to the Excel file in the "master" folder
EXCEL_FILE_PATH = os.path.join("master", "NSE_symbol_main_local.xlsx")

# Ensure the master directory exists
os.makedirs("master", exist_ok=True)

# Load or create the Excel workbook
if os.path.exists(EXCEL_FILE_PATH):
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
else:
    workbook = openpyxl.Workbook()
    workbook.save(EXCEL_FILE_PATH)

# Check and define the Excel sheet name
if base_sheet_name in workbook.sheetnames:
    excel_sheet_name = f"{base_sheet_name}_{ist_now.strftime('%H-%M-%S')}"
else:
    excel_sheet_name = base_sheet_name

# Create or get the Excel sheet
worksheet_excel = workbook.create_sheet(title=excel_sheet_name)
worksheet_excel.append(['Symbol'] + headers)

# Process each symbol
processed_count = 0
for symbol in symbols:
    fetch_and_update_stock_data(symbol, new_worksheet, worksheet_excel)
    processed_count += 1
    # Add a delay to avoid rate-limiting
    time.sleep(1)
    log_message(f"Processed {processed_count}/{len(symbols)} symbols.")

# Save the updated Excel file
workbook.save(EXCEL_FILE_PATH)
log_message(f"Excel file updated and saved at {EXCEL_FILE_PATH}.")

# Commit the updated Excel file to Git
def git_commit(file_path):
    try:
        subprocess.run(['git', 'add', file_path], check=True)
        subprocess.run(['git', 'commit', '-m', f"Added {file_path}"], check=True)
        subprocess.run(['git', 'push'], check=True)
        log_message(f"Successfully committed {file_path} to Git.")
    except subprocess.CalledProcessError as e:
        log_message(f"Error committing {file_path} to Git: {e}")

git_commit(LOG_FILE_PATH)
git_commit(CSV_FILE_PATH)
git_commit(EXCEL_FILE_PATH)

log_message(f"Log saved to {LOG_FILE_PATH}.")
log_message(f"CSV data saved to {CSV_FILE_PATH}.")
log_message(f"Excel data saved to {EXCEL_FILE_PATH}.")

