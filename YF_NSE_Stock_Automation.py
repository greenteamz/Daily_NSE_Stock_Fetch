import yfinance as yf
import pandas as pd
from datetime import date, datetime
import os
from openpyxl import Workbook, load_workbook
import zipfile

# List all the Excel files in the current directory
excel_files = [f for f in os.listdir() if f.endswith('.xlsx')]

# If there are no Excel files, print a message and exit
if not excel_files:
    print("No Excel files found in the directory.")
    exit()

# Print all Excel files found in the directory
print("Available Excel files:")
for i, file in enumerate(excel_files, start=1):
    print(f"{i}. {file}")

# Ask the user to choose an Excel file from the list
try:
    choice = int(input(f"Enter the number corresponding to the Excel file you want to use (1-{len(excel_files)}): "))
    selected_file = excel_files[choice - 1]
except (ValueError, IndexError):
    print("Invalid choice. Exiting the program.")
    exit()

# Load the selected Excel file using openpyxl
try:
    wb = load_workbook(selected_file)
except zipfile.BadZipFile:
    print(f"Error: {selected_file} is not a valid Excel file. It may be corrupted.")
    exit()

# Access the worksheet (tab) named 'symbol' in the workbook, or create it if it doesn't exist
if 'symbol' in wb.sheetnames:
    # Read 'symbol' sheet into a DataFrame
    source_worksheet = pd.read_excel(selected_file, sheet_name='symbol')
else:
    # Create a new 'symbol' sheet if it doesn't exist
    source_worksheet = pd.DataFrame(columns=['Symbol'])
    # Write a new 'symbol' sheet
    source_worksheet.to_excel(selected_file, sheet_name='symbol', index=False)

# Fetch all stock symbols from the first column
symbols = source_worksheet['Symbol'].dropna().tolist()  # Drop NaN values and get the list of symbols

# Print the total number of symbols
total_symbols = len(symbols)
print(f"Total number of symbols: {total_symbols}")

# Add '.NS' suffix to all symbols
symbols = [symbol if symbol.endswith('.NS') else f"{symbol}.NS" for symbol in symbols]

# Get today's date for the new sheet name
today = date.today().strftime('%d_%m_%Y')
sheet_base_name = f"Yearly_Data_{today}"

# Check if a sheet with the same name already exists
if sheet_base_name in wb.sheetnames:
    current_time = datetime.now().strftime('%H_%M')
    sheet_name = f"{sheet_base_name}__{current_time}"
else:
    sheet_name = sheet_base_name

# Add a new sheet for yearly data
sheet = wb.create_sheet(sheet_name)

# Write the headers to the new sheet
headers = ['Date', 'Open', 'High', 'Low', 'Close', 'Adj Close', 'Volume', 'Ticker']
sheet.append(headers)

# Fetch and write data for each symbol
try:
    for symbol in symbols:
        print(f"Fetching data for {symbol}...")
        data = yf.download(tickers=symbol, period="max", interval="1d", auto_adjust=True)
        if data.empty:
            print(f"No data found for {symbol}. Skipping...")
            continue

        # Strip timezone info from the Date column
        data.index = data.index.tz_localize(None)

        # Add a column for the ticker symbol
        data['Ticker'] = symbol
        data.reset_index(inplace=True)

        # Write the data to the new sheet
        for _, row in data.iterrows():
            sheet.append(row.tolist())

        # Save the workbook incrementally
        wb.save(selected_file)

    print(f"Yearly data successfully added to the new worksheet titled '{sheet_name}'.")
except Exception as e:
    print(f"Error fetching data: {e}")

# Ensure all temporary files are closed properly
finally:
    wb.close()
